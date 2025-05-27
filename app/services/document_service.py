import asyncio
import json
import logging
import os
import pickle
import re
import shutil
import tempfile
import time
import uuid
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union
from urllib.parse import unquote, urlparse

import aiohttp
import chromadb
import docx2txt
import faiss
import numpy as np
import openpyxl  # Added import for openpyxl
from dotenv import load_dotenv
from langchain_text_splitters import RecursiveCharacterTextSplitter
from pypdf import PdfReader
from rank_bm25 import BM25Okapi

from app.infra.chroma_wrapper import ChromaEmbeddingWrapper
from app.utils.chunkers import chunk_document

# Load environment variables
load_dotenv()

# Configure logger
logger = logging.getLogger("cx_consulting_ai.document_service")

# Define collection name constants / patterns
CX_GLOBAL_COLLECTION = "global_kb"
DELIVERABLE_COLLECTION_PREFIX = "deliverable_"
USER_COLLECTION_PREFIX = "user_"


class DocumentService:
    """Service for document processing and retrieval."""

    def __init__(
        self,
        documents_dir: str,
        chunked_dir: str,
        vectorstore_dir: str,
        embedding_model: Optional[str] = None,
        chunk_size: Optional[int] = None,
        chunk_overlap: Optional[int] = None,
        default_collection_name: str = "cx_documents",
    ):
        """
        Initialize the document service.

        Args:
            documents_dir: Directory containing source documents
            chunked_dir: Directory for storing chunked documents
            vectorstore_dir: Directory for the vector database
            embedding_model: Embedding model name
            chunk_size: Maximum chunk size
            chunk_overlap: Chunk overlap size
            default_collection_name: Default Chroma collection name
        """
        # Set directories
        self.documents_dir = documents_dir
        self.chunked_dir = chunked_dir
        self.vectorstore_dir = vectorstore_dir
        self.default_collection_name = default_collection_name
        self.active_collection_name: Optional[str] = None
        self.chroma_client: Optional[chromadb.ClientAPI] = None
        self.collection: Optional[chromadb.Collection] = None
        self.collection_cache: Dict[str, chromadb.Collection] = {}
        self.embedding_model_name_constructor_arg = embedding_model
        self.embedding_model_to_use_for_init = embedding_model
        self.embedding_manager = None
        self._doc_embedding_fn = None
        self._query_embedding_fn = None
        self.chroma_embedding_function = None
        self.bm25_indices: Dict[str, BM25Okapi] = {}
        self.bm25_corpus: Dict[str, List[str]] = {}
        self.bm25_doc_ids: Dict[str, List[str]] = {}
        self.bm25_storage_dir = Path(self.vectorstore_dir) / "bm25_data"
        os.makedirs(self.bm25_storage_dir, exist_ok=True)

        # Create directories if they don't exist
        os.makedirs(self.documents_dir, exist_ok=True)
        os.makedirs(self.chunked_dir, exist_ok=True)
        os.makedirs(self.vectorstore_dir, exist_ok=True)

        # Set embedding model and chunking parameters
        # The self.embedding_model will be None if not provided by constructor.
        # _init_embeddings will then pick it up from settings.
        self.embedding_model = embedding_model

        self.chunk_size = chunk_size or int(os.getenv("MAX_CHUNK_SIZE", "512"))
        self.chunk_overlap = chunk_overlap or int(os.getenv("CHUNK_OVERLAP", "50"))

        # Vector database type (Chroma or FAISS)
        self.vector_db_type = os.getenv("VECTOR_DB_TYPE", "chroma").lower()

        # FAISS specific paths, defined here for clarity even if FAISS not used
        self.faiss_index_path = str(Path(self.vectorstore_dir) / "faiss_index.bin")
        self.faiss_docstore_path = str(
            Path(self.vectorstore_dir) / "faiss_docstore.pkl"
        )  # Changed to .pkl for pickle

        self._init_embeddings()
        # self._init_vector_store() # This is now async, called from main.py startup

        logger.info(
            f"Document Service instance created (vector_db_type: {self.vector_db_type}, default_collection: '{self.default_collection_name}'). Async _init_vector_store must be called explicitly."
        )

    def _get_collection_name_for_doc(self, doc_metadata: Dict[str, Any]) -> str:
        """Determine the target collection name based on document metadata."""
        if doc_metadata.get("is_global", False):
            logger.debug(
                f"Routing document (source: {doc_metadata.get('source')}) to GLOBAL collection."
            )
            return CX_GLOBAL_COLLECTION

        # Check for deliverable templates (requires template_name in metadata - see task V-2)
        # For now, we can check document_type or add a placeholder logic
        doc_type = doc_metadata.get("document_type", "")
        template_name = doc_metadata.get("template_name")  # Task V-2 will add this
        if template_name:  # Check if template_name is present
            # Sanitize template_name for collection name
            sanitized_template_name = re.sub(r"[^a-zA-Z0-9_-]+", "_", template_name)
            collection_name = (
                f"{DELIVERABLE_COLLECTION_PREFIX}{sanitized_template_name}"
            )
            # Chroma limits: 3-63 chars, start/end alphanumeric, no .. or consecutive .
            collection_name = collection_name[:60]  # Truncate
            # Basic cleanup, might need more sophisticated validation for Chroma names
            if len(collection_name) < 3:
                collection_name = f"{collection_name}___"  # Pad if too short
            logger.debug(
                f"Routing document (source: {doc_metadata.get('source')}) to DELIVERABLE collection: {collection_name}"
            )
            return collection_name

        project_id = doc_metadata.get("project_id")
        if project_id:
            # Sanitize project_id for collection name
            sanitized_project_id = re.sub(r"[^a-zA-Z0-9_-]+", "_", project_id)
            collection_name = f"{USER_COLLECTION_PREFIX}{sanitized_project_id}"
            collection_name = collection_name[:60]  # Truncate
            if len(collection_name) < 3:
                collection_name = f"{collection_name}___"  # Pad
            logger.debug(
                f"Routing document (source: {doc_metadata.get('source')}) to USER collection: {collection_name}"
            )
            return collection_name

        # Fallback to default
        logger.debug(
            f"Routing document (source: {doc_metadata.get('source')}) to DEFAULT collection: {self.default_collection_name}"
        )
        return self.default_collection_name

    def _resolve_embedding_model_name(
        self, preferred_name: Optional[str] = None
    ) -> Optional[str]:
        """Resolves the effective embedding model name based on priority."""
        from app.core.config import get_settings  # Local import for settings access

        settings = get_settings()

        if preferred_name:
            return preferred_name
        if self.embedding_model_name_constructor_arg:
            return self.embedding_model_name_constructor_arg

        # Fallback to settings if no specific name is given
        if settings.EMBEDDING_TYPE.lower() == "bge":
            return settings.BGE_MODEL_NAME
        else:
            return settings.EMBEDDING_MODEL

    def _init_embeddings(self) -> None:
        """Initialize the embedding manager and functions based on self.embedding_model_to_use_for_init."""
        try:
            from app.core.config import get_settings
            from app.utils.embedding_manager import create_embedding_manager

            settings = get_settings()

            effective_model_name = self._resolve_embedding_model_name(
                self.embedding_model_to_use_for_init
            )
            if not effective_model_name:
                raise ValueError(
                    "Effective embedding model name could not be resolved for _init_embeddings."
                )

            logger.info(f"Initializing embeddings with model: {effective_model_name}")
            self.embedding_manager = create_embedding_manager(
                model_type=settings.EMBEDDING_TYPE.lower(),
                model_name=effective_model_name,
                device=settings.EMBEDDING_DEVICE,
                use_offline_fallback=True,
            )
            self._doc_embedding_fn = self.embedding_manager.get_embedding_function(
                for_queries=False
            )
            self._query_embedding_fn = self.embedding_manager.get_embedding_function(
                for_queries=True
            )
            self.chroma_embedding_function = ChromaEmbeddingWrapper(
                self.embedding_manager
            )
            self.embedding_model_to_use_for_init = (
                effective_model_name  # Store the name actually used
            )
            logger.info(
                f"Embeddings initialized: Type={self.embedding_manager.model_type}, "
                f"Model={self.embedding_manager.model_name}, Dim={self.embedding_manager.dimension_size}"
            )
        except Exception as e:
            logger.error(f"Error initializing embeddings: {e}", exc_info=True)
            self._init_local_fallback()  # Fallback if actual init fails

    def _ensure_embeddings_initialized(
        self, requested_model_name: Optional[str] = None
    ) -> None:
        """Ensures embeddings are initialized, re-initializing if model name changes."""
        current_resolved_name_for_init = self._resolve_embedding_model_name(
            self.embedding_model_to_use_for_init
        )
        target_model_name = self._resolve_embedding_model_name(requested_model_name)

        if not self.embedding_manager or (
            target_model_name and current_resolved_name_for_init != target_model_name
        ):
            logger.info(
                f"Re-evaluating embedding initialization. Requested: {target_model_name}, Current used for init: {current_resolved_name_for_init}"
            )
            self.embedding_model_to_use_for_init = (
                target_model_name  # Set what _init_embeddings should use
            )
            self._init_embeddings()
        elif (
            not self.chroma_embedding_function
        ):  # Catch if manager is there but chroma_wrapper isn't
            logger.warning(
                "ChromaEmbeddingWrapper is missing despite EmbeddingManager presence. Re-initializing embeddings."
            )
            self._init_embeddings()

    def _init_local_fallback(self):
        """Initialize with a simple local embedding function as fallback."""
        import hashlib

        logger.info("Using simple local embedding function (offline mode)")

        # Define a simple embedding function using hash values
        def simple_embedding(text):
            if isinstance(text, list):
                return [simple_embedding(t) for t in text]

            # Create a hash of the text
            hash_object = hashlib.md5(text.encode())
            hash_hex = hash_object.hexdigest()

            # Convert hash to a vector of floats
            vector = []
            for i in range(0, len(hash_hex), 2):
                if i + 2 <= len(hash_hex):
                    hex_pair = hash_hex[i : i + 2]
                    value = int(hex_pair, 16) / 255.0  # Normalize to [0, 1]
                    vector.append(value)

            # Pad to 384 dimensions for a reasonably sized vector
            vector = vector + [0.0] * (384 - len(vector))
            return vector

        # Set embedding functions
        self._doc_embedding_fn = lambda texts: simple_embedding(texts)
        self._query_embedding_fn = lambda text: [
            simple_embedding(text)
        ]  # Keep List[List[float]] for consistency for now

        logger.info("Local embedding function initialized")

    def _init_chroma(self):
        """Initializes ChromaDB client and related configurations. Synchronous."""
        if not self.chroma_embedding_function:
            # This should have been set by _init_embeddings() called in __init__
            logger.error(
                "Chroma embedding function wrapper not initialized prior to _init_chroma. This indicates an issue in initialization order."
            )
            # Attempt to re-initialize embeddings, though this is a recovery attempt.
            self._init_embeddings()
            if not self.chroma_embedding_function:
                raise RuntimeError(
                    "Failed to initialize Chroma embedding function even after recovery attempt."
                )

        try:
            logger.info(
                f"Initializing ChromaDB persistent client at path: {self.vectorstore_dir}"
            )
            self.chroma_client = chromadb.PersistentClient(path=self.vectorstore_dir)
            logger.info("ChromaDB client initialized successfully.")
            # Default collection will be set by _init_vector_store -> set_collection
        except Exception as e:
            logger.error(f"Failed to initialize ChromaDB client: {e}", exc_info=True)
            raise RuntimeError("ChromaDB client initialization failed") from e

    def _init_faiss(self):
        """Initializes FAISS index and docstore. Synchronous."""
        logger.info(
            f"Initializing FAISS store. Index path: {self.faiss_index_path}, Docstore path: {self.faiss_docstore_path}"
        )
        if (
            not hasattr(self, "embedding_manager")
            or not self.embedding_manager.dimension_size
        ):
            logger.error(
                "Embedding manager not initialized or dimension size unknown. Cannot initialize FAISS."
            )
            # Attempt to re-initialize embeddings as a recovery measure.
            self._init_embeddings()
            if (
                not hasattr(self, "embedding_manager")
                or not self.embedding_manager.dimension_size
            ):
                raise RuntimeError(
                    "FAISS initialization failed: Embedding dimension unknown even after recovery."
                )

        embedding_dim = self.embedding_manager.dimension_size

        try:
            if os.path.exists(self.faiss_index_path) and os.path.exists(
                self.faiss_docstore_path
            ):
                logger.info("Loading existing FAISS index and docstore...")
                self.index = faiss.read_index(self.faiss_index_path)
                with open(self.faiss_docstore_path, "rb") as f:
                    self.docstore = pickle.load(f)

                if self.index.d != embedding_dim:
                    logger.warning(
                        f"Loaded FAISS index dimension ({self.index.d}) does not match current model dimension ({embedding_dim}). Re-initializing FAISS store."
                    )
                    self.index = faiss.IndexFlatL2(embedding_dim)
                    self.docstore = {"ids": [], "documents": []}
                else:
                    logger.info(
                        f"Loaded FAISS index with {self.index.ntotal} vectors and dimension {self.index.d}."
                    )
            else:
                logger.info("No existing FAISS index found. Creating new FAISS store.")
                self.index = faiss.IndexFlatL2(embedding_dim)
                self.docstore = {"ids": [], "documents": []}
            logger.info("FAISS store initialized/loaded successfully.")
        except Exception as e:
            logger.error(f"Error initializing FAISS store: {e}", exc_info=True)
            # Fallback to empty store on error
            logger.warning(
                "Falling back to empty FAISS store due to initialization error."
            )
            self.index = faiss.IndexFlatL2(embedding_dim)
            self.docstore = {"ids": [], "documents": []}
            # Do not raise here, allow service to start with an empty FAISS store if loading failed.

    async def _init_vector_store(self):
        """Initialize the vector store with the default collection. Async."""
        logger.debug("Initializing vector store (async)...")
        self._ensure_embeddings_initialized()  # Ensures embeddings are ready
        if self.vector_db_type == "chroma":
            self._init_chroma()  # This is synchronous
            if self.default_collection_name:
                try:
                    logger.info(
                        f"Attempting to initialize default Chroma collection (async): {self.default_collection_name}"
                    )
                    await self.set_collection(self.default_collection_name)
                except Exception as e:
                    logger.error(
                        f"Failed to initialize default Chroma collection '{self.default_collection_name}' during async init: {e}",
                        exc_info=True,
                    )
        elif self.vector_db_type == "faiss":
            self._init_faiss()  # This is synchronous
        else:
            logger.warning(
                f"Unknown vector_db_type: {self.vector_db_type}. Vector store not fully initialized."
            )
        logger.debug(
            f"Vector store initialized using {self.vector_db_type}. Chroma client: {self.chroma_client is not None}"
        )

    async def set_collection(
        self,
        collection_name: str,
        force_reload: bool = False,
        new_embedding_model_name: Optional[str] = None,
    ):
        """Sets the active collection. Re-initializes embeddings if new_embedding_model_name is different."""
        logger.debug(
            f"Set collection: '{collection_name}', force_reload={force_reload}, new_model: {new_embedding_model_name}"
        )
        self._ensure_embeddings_initialized(new_embedding_model_name)  # P2 Fix

        self.active_collection_name = collection_name

        # Load BM25 data if not already loaded for this collection
        if (
            collection_name not in self.bm25_indices
        ):  # Check if index is missing as a proxy for all BM25 data
            logger.info(
                f"BM25 data for '{collection_name}' not in memory. Attempting to load from disk."
            )
            await self._load_bm25_data(collection_name)
        else:
            logger.debug(f"BM25 data for '{collection_name}' already in memory.")

        if self.vector_db_type == "chroma":
            if not self.chroma_client:
                logger.error("Chroma client not initialized for set_collection.")
                raise RuntimeError("Chroma client not available.")

            if collection_name in self.collection_cache:
                self.collection = self.collection_cache[collection_name]
                logger.debug(f"Using cached Chroma collection: '{collection_name}'")
                return

            if not self.chroma_embedding_function:
                logger.error(
                    "Chroma embedding function not available for set_collection after ensuring init."
                )
                raise RuntimeError(
                    "Chroma embedding function could not be initialized."
                )

            try:
                logger.info(
                    f"Getting or creating Chroma collection '{collection_name}' with embedding model '{self.embedding_manager.model_name if self.embedding_manager else 'N/A'}'."
                )

                def _get_or_create_sync():
                    return self.chroma_client.get_or_create_collection(
                        name=collection_name,
                        embedding_function=self.chroma_embedding_function,  # Crucial: use the (re)initialized one
                    )

                collection = await asyncio.to_thread(_get_or_create_sync)
                self.collection = collection
                self.collection_cache[collection_name] = collection
                logger.info(
                    f"Active Chroma collection is now: '{self.collection.name}'"
                )

            except Exception as e:
                logger.error(
                    f"Failed to set Chroma collection to '{collection_name}': {e}",
                    exc_info=True,
                )
                self.collection = None
                raise RuntimeError(
                    f"Failed to set Chroma collection '{collection_name}'"
                ) from e
        elif self.vector_db_type == "faiss":
            # For FAISS, self.active_collection_name is set. If multiple FAISS dbs were managed by name,
            # logic to load the correct one would go here.
            logger.debug(
                f"FAISS mode: active collection name set to '{collection_name}'. Ensure corresponding FAISS index is loaded if applicable."
            )
        else:
            logger.warning(
                f"Setting collection for unsupported vector_db_type: {self.vector_db_type}"
            )

    async def add_document(
        self,
        document_url: str,
        document_type: str,
        metadata: Optional[Dict[str, Any]] = None,
        is_global: bool = False,
        project_id: Optional[str] = None,
        collection_name: Optional[str] = None,
    ) -> bool:
        """
        Add a document to the vector store asynchronously.
        Uses collection policy based on metadata. Prepares chunk metadata & IDs, then adds chunks.
        """
        # Prepare metadata first
        doc_metadata = metadata.copy() if metadata else {}
        doc_metadata.setdefault("source", document_url)
        doc_metadata.setdefault(
            "project_id", project_id
        )  # Ensure project_id from args is in meta
        doc_metadata.setdefault(
            "is_global", is_global
        )  # Ensure is_global from args is in meta
        doc_metadata.setdefault("document_type", document_type)
        doc_metadata.setdefault("processed_at", time.time())
        doc_metadata.setdefault(
            "doc_id", f"{Path(document_url).stem}_{int(time.time())}"
        )
        # Add other metadata if available (e.g., template_name - needed for policy)
        # Assuming 'metadata' arg might contain 'template_name'
        if metadata and "template_name" in metadata:
            doc_metadata.setdefault("template_name", metadata["template_name"])

        # Determine target collection using the policy
        target_collection_name = self._get_collection_name_for_doc(doc_metadata)

        # Allow explicit override via collection_name argument IF PROVIDED
        # (Consider removing this override later if policy should be strict)
        if collection_name:
            logger.warning(
                f"Overriding collection policy for {document_url}. Explicitly using collection: {collection_name}"
            )
            target_collection_name = collection_name

        if not target_collection_name:
            logger.error(
                "Collection policy failed to determine a target collection name."
            )
            return False

        with tempfile.TemporaryDirectory(prefix="doc_processing_") as temp_dir_str:
            temp_dir_path = Path(temp_dir_str)
            downloaded_file_path: Optional[Path] = (
                None  # To store the path of the successfully downloaded file
            )

            try:
                logger.info(
                    f"Adding document: {document_url}, Type: {document_type}, Target Collection: {target_collection_name}"
                )
                # doc_metadata is already populated above

                needs_to_set_collection = (
                    self.active_collection_name != target_collection_name
                ) or (
                    self.vector_db_type == "chroma"
                    and (
                        not self.collection
                        or self.collection.name != target_collection_name
                    )
                )
                if needs_to_set_collection:
                    logger.info(
                        f"Setting active collection to '{target_collection_name}' for document addition."
                    )
                    await self.set_collection(target_collection_name)

                # Download document - now pass the temp_dir_path to it
                downloaded_file_path = await self._download_document(
                    document_url, document_type, temp_dir_path
                )
                if not downloaded_file_path:
                    return False  # Error logged by _download_document

                # Extract text (this now returns text and handles cleanup of downloaded_file_path)
                # Note: _extract_text needs modification to NOT delete file_path if chunker needs it
                # Let's assume _extract_text returns the text and doesn't delete.
                # We will need to modify _extract_text next or handle cleanup differently.
                # FOR NOW: Assume file_path remains after _extract_text call

                # --- Use new chunker ---
                # Pass the path to the downloaded file and original metadata
                chunk_data = await asyncio.to_thread(  # Run sync chunker in thread
                    chunk_document,
                    file_path=downloaded_file_path,  # Use the path of the file in the temp dir
                    file_type=document_type,
                    doc_metadata=doc_metadata,
                )

                if not chunk_data:
                    logger.warning(
                        f"No chunks created for document {doc_metadata['doc_id']} from {document_url} using new chunker."
                    )
                    # No explicit cleanup of downloaded_file_path needed here, TemporaryDirectory handles it
                    return False

                # Prepare lists for vector store addition
                chunks_texts = [item["text"] for item in chunk_data]
                metadatas_list = [item["metadata"] for item in chunk_data]
                # Generate IDs based on original doc_id and chunk index
                ids_list = [
                    f"{doc_metadata['doc_id']}_chunk_{i}"
                    for i in range(len(chunk_data))
                ]
                # Add chunk_id and preview to metadata AFTER getting it from chunker
                for i, meta in enumerate(metadatas_list):
                    meta["chunk_id"] = ids_list[i]
                    meta["chunk_index"] = i
                    meta["chunk_text_preview"] = chunks_texts[i][:100] + (
                        "..." if len(chunks_texts[i]) > 100 else ""
                    )

                # Add to vector store
                await self._add_chunks_to_vectorstore(
                    chunks=chunks_texts,
                    metadatas_list=metadatas_list,
                    ids_list=ids_list,
                    source_uri=doc_metadata.get("source"),
                    project_id=doc_metadata.get("project_id"),
                    is_global=doc_metadata.get("is_global", False),
                )

                logger.info(
                    f"Successfully processed and initiated addition of {len(chunks_texts)} chunks from {document_url} to collection '{self.active_collection_name}' with doc_id '{doc_metadata['doc_id']}'"
                )
                return True

            except Exception as e_exc:
                logger.error(
                    f"Error adding document {document_url}: {e_exc}", exc_info=True
                )
                return False

    async def _download_document(
        self,
        document_url: str,
        document_type: str,
        target_temp_dir: Path,  # Added target_temp_dir
    ) -> Optional[Path]:
        """Download a document from a URL or copy a local file to a temporary location within target_temp_dir."""
        # Define a temporary directory within the main document storage area
        # temp_dir = Path(self.documents_dir) / "_temp_downloads" # No longer create its own subdir like this
        # temp_dir.mkdir(parents=True, exist_ok=True) # No longer needed

        try:
            parsed_url = urlparse(document_url)
            scheme = parsed_url.scheme.lower()
            is_local = scheme == "" or scheme == "file"
            is_remote = scheme in [
                "http",
                "https",
                "s3",
            ]  # Add other remote schemes if needed

            if is_remote:
                # Generate a unique temporary filename
                temp_filename = f"{uuid.uuid4()}.{document_type}"
                temp_filepath = target_temp_dir / temp_filename

                logger.info(
                    f"Downloading remote URL: {document_url} to {temp_filepath}"
                )
                async with aiohttp.ClientSession() as session:
                    async with session.get(document_url) as response:
                        response.raise_for_status()  # Raise an exception for bad status codes (4xx or 5xx)
                        with open(temp_filepath, "wb") as f:
                            while True:
                                chunk = await response.content.read(
                                    1024
                                )  # Read in chunks
                                if not chunk:
                                    break
                                f.write(chunk)
                logger.info(f"Downloaded successfully: {document_url}")
                return temp_filepath

            elif is_local:
                # Handle local file path (could be absolute or relative)
                # Resolve the path relative to CWD if it's not absolute
                local_path = Path(
                    unquote(parsed_url.path)
                )  # Decode URL-encoded characters like %20
                if not local_path.is_absolute():
                    local_path = Path.cwd() / local_path
                    logger.debug(
                        f"Resolved relative local path '{unquote(parsed_url.path)}' to '{local_path}'"
                    )
                else:
                    logger.debug(f"Using absolute local path: {local_path}")

                if not local_path.exists():
                    logger.error(f"Local file not found: {local_path}")
                    raise FileNotFoundError(f"Local file not found: {local_path}")
                if local_path.is_dir():
                    logger.error(f"Local path is a directory, not a file: {local_path}")
                    raise IsADirectoryError(f"Path is a directory: {local_path}")

                # Generate a unique temporary filename using original extension if possible
                original_suffix = local_path.suffix or f".{document_type}"
                temp_filename = f"{uuid.uuid4()}{original_suffix}"
                temp_filepath = (
                    target_temp_dir / temp_filename
                )  # Use passed target_temp_dir

                logger.info(f"Copying local file: {local_path} to {temp_filepath}")
                shutil.copy2(local_path, temp_filepath)  # copy2 preserves metadata
                logger.info(f"Copied successfully: {local_path}")
                return temp_filepath

            else:
                logger.error(
                    f"Unsupported URL scheme '{scheme}' for document: {document_url}"
                )
                return None

        except FileNotFoundError as fnf_error:
            logger.error(
                f"File not found error during processing: {fnf_error}", exc_info=True
            )
            return None  # Or re-raise if needed
        except IsADirectoryError as dir_error:
            logger.error(
                f"Path is directory error during processing: {dir_error}", exc_info=True
            )
            return None
        except aiohttp.ClientError as http_error:
            logger.error(
                f"HTTP error downloading document {document_url}: {http_error}",
                exc_info=True,
            )
            return None
        except Exception as e:
            logger.error(
                f"Error processing document URL {document_url}: {e}", exc_info=True
            )
            return None

    async def _extract_text(self, file_path: Path, document_type: str) -> Optional[str]:
        file_path_obj = Path(file_path)
        logger.info(f"Extracting text from {file_path_obj} (type: {document_type})")
        extracted_content: Optional[str] = None

        try:
            if document_type == "pdf":
                text = ""
                try:
                    reader = PdfReader(str(file_path_obj))
                    for page in reader.pages:
                        text += page.extract_text() or ""
                    logger.info(
                        f"Extracted {len(text)} chars from PDF: {file_path_obj.name}"
                    )
                    if not text.strip():
                        logger.warning(
                            f"No text extracted from PDF (possibly image-based or empty): {file_path_obj.name}"
                        )
                        try:
                            import pytesseract
                            from pdf2image import convert_from_path

                            logger.info(f"Attempting OCR on PDF: {file_path_obj.name}")
                            images = convert_from_path(str(file_path_obj))
                            ocr_text = ""
                            for i, img in enumerate(images):
                                ocr_text += pytesseract.image_to_string(img) + "\n\n"
                            text = ocr_text
                        except (ImportError, Exception) as ocr_err:
                            logger.error(
                                f"OCR attempt failed on PDF {file_path_obj.name}: {ocr_err}",
                                exc_info=True,
                            )

                except ImportError:
                    logger.warning(
                        f"pdf2image or pytesseract not installed. Cannot perform OCR on PDF: {file_path_obj.name}"
                    )
                except Exception as ocr_err:
                    logger.error(
                        f"Error during OCR on PDF {file_path_obj.name}: {ocr_err}",
                        exc_info=True,
                    )
                except Exception as pdf_err:
                    logger.error(
                        f"Error reading PDF {file_path_obj.name}: {pdf_err}",
                        exc_info=True,
                    )
                    # Keep text as potentially empty from OCR attempt or previous state
                extracted_content = text  # Assign text potentially modified by OCR
            elif document_type == "txt":
                with open(file_path_obj, "r", encoding="utf-8") as f:
                    extracted_content = f.read()
            elif document_type == "md":
                with open(file_path_obj, "r", encoding="utf-8") as f:
                    extracted_content = f.read()
            elif document_type == "docx":
                try:
                    extracted_content = docx2txt.process(str(file_path_obj))
                except Exception as docx_err:
                    logger.error(
                        f"Error processing DOCX file {file_path_obj.name}: {docx_err}",
                        exc_info=True,
                    )
                    extracted_content = None  # Set to None on error
            elif document_type == "csv":
                try:
                    import csv  # Ensure csv module is imported

                    text_parts = []
                    with open(file_path_obj, "r", encoding="utf-8", newline="") as f:
                        reader = csv.reader(f)
                        for row in reader:
                            text_parts.append(", ".join(row))
                    extracted_content = "\\n".join(text_parts)
                    logger.info(
                        f"Extracted {len(extracted_content)} chars from CSV: {file_path_obj.name}"
                    )
                except Exception as csv_err:
                    logger.error(
                        f"Error processing CSV file {file_path_obj.name}: {csv_err}",
                        exc_info=True,
                    )
                    extracted_content = None
            elif document_type == "xlsx":
                try:
                    import openpyxl  # Ensure openpyxl is imported

                    workbook = openpyxl.load_workbook(
                        file_path_obj, data_only=True
                    )  # data_only=True to get values, not formulas
                    text_parts = []
                    for sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        for row in sheet.iter_rows():
                            row_texts = [
                                str(cell.value) if cell.value is not None else ""
                                for cell in row
                            ]
                            text_parts.append(
                                ", ".join(filter(None, row_texts))
                            )  # Join non-empty cell values
                    extracted_content = "\\n".join(
                        filter(None, text_parts)
                    )  # Join non-empty rows
                    logger.info(
                        f"Extracted {len(extracted_content)} chars from XLSX: {file_path_obj.name}"
                    )
                except ImportError:
                    logger.error(
                        "openpyxl library not installed. Please install it to process .xlsx files."
                    )
                    extracted_content = None
                except Exception as xlsx_err:
                    logger.error(
                        f"Error processing XLSX file {file_path_obj.name}: {xlsx_err}",
                        exc_info=True,
                    )
                    extracted_content = None
            elif document_type == "json":
                try:
                    with open(file_path_obj, "r", encoding="utf-8") as f:
                        data = json.load(f)
                    extracted_content = json.dumps(data, indent=2)
                except json.JSONDecodeError as json_err:
                    logger.error(
                        f"Error decoding JSON file {file_path_obj.name}: {json_err}",
                        exc_info=True,
                    )
                    extracted_content = None
                except Exception as json_read_err:
                    logger.error(
                        f"Error reading JSON file {file_path_obj.name}: {json_read_err}",
                        exc_info=True,
                    )
                    extracted_content = None
            else:
                logger.warning(f"Unsupported document type: {document_type}")
                extracted_content = None
        except Exception as e:
            logger.error(
                f"Unexpected error during text extraction setup for {file_path_obj.name}: {e}",
                exc_info=True,
            )
            extracted_content = None

        return extracted_content  # Return the content extracted (or None)

    async def _chunk_text(
        self, text: str, document_id: str, metadata: Dict[str, Any]
    ) -> List[Dict[str, Any]]:
        if not text or not text.strip():
            logger.warning(f"Cannot chunk empty text for document_id: {document_id}")
            return []

        chunks = []  # Initialize chunks
        try:
            text_splitter = RecursiveCharacterTextSplitter(
                chunk_size=self.chunk_size,
                chunk_overlap=self.chunk_overlap,
                length_function=len,
                is_separator_regex=False,
            )
            chunks = text_splitter.split_text(text)
        except Exception as split_err:
            logger.error(
                f"Error splitting text for document {document_id}: {split_err}",
                exc_info=True,
            )
            return []  # Return empty list on error

        # This part is outside the try/except for splitting
        chunk_documents = []
        for i, chunk_text in enumerate(chunks):  # Iterate over the split chunks
            chunk_id = f"{document_id}_chunk_{i}"
            chunk_metadata = metadata.copy()
            chunk_metadata["chunk_id"] = chunk_id
            chunk_metadata["chunk_index"] = i
            chunk_documents.append(
                {"text": chunk_text, "id": chunk_id, "metadata": chunk_metadata}
            )
        return chunk_documents

    async def _add_chunks_to_vectorstore(
        self,
        chunks: List[str],
        metadatas_list: List[Dict[str, Any]],
        ids_list: List[str],
        source_uri: Optional[str] = None,
        project_id: Optional[str] = None,
        is_global: bool = False,
    ) -> None:
        """Adds text chunks and their embeddings to the vector store (Chroma or FAISS) and updates BM25 index asynchronously."""
        if not chunks:
            logger.debug("No chunks provided to _add_chunks_to_vectorstore. Skipping.")
            return

        if not self.active_collection_name:
            logger.error(
                "No active collection selected for adding chunks. Call set_collection() first."
            )
            return

        # --- Fallback check for embedding function ---
        if not self._doc_embedding_fn:
            logger.warning(
                "Embedding function (self._doc_embedding_fn) is not initialized. Attempting fallback."
            )
            self._ensure_embeddings_initialized()  # Attempt to initialize fully first
            if not self._doc_embedding_fn:  # Check again
                logger.error(
                    "Full embedding initialization failed after re-attempt. Cannot add chunks without embedding function."
                )
                return  # Exit if embedding function is still unavailable

        logger.info(
            f"Adding {len(chunks)} chunks to '{self.active_collection_name}' (DB: {self.vector_db_type})."
        )
        start_time = time.time()

        # The self._doc_embedding_fn should be the actual callable function
        try:
            # Ensure _doc_embedding_fn is called in a thread as it's synchronous
            if asyncio.iscoroutinefunction(self._doc_embedding_fn):
                # This branch should ideally not be hit if _doc_embedding_fn is always sync from EmbeddingManager
                logger.warning(
                    "_doc_embedding_fn appears to be a coroutine function. Awaiting directly."
                )
                embeddings = await self._doc_embedding_fn(chunks)
            else:
                # Run synchronous function in thread
                loop = asyncio.get_running_loop()
                embeddings = await loop.run_in_executor(
                    None, self._doc_embedding_fn, chunks
                )
        except Exception as e:
            logger.error(
                f"Error generating chunk embeddings for '{self.active_collection_name}': {e}",
                exc_info=True,
            )
            return

        if not embeddings or len(embeddings) != len(chunks):
            logger.error(
                f"Failed to generate embeddings or mismatch in count for '{self.active_collection_name}'. Expected {len(chunks)}, got {len(embeddings) if embeddings else 0}."
            )
            return

        # Prepare metadatas (as it was, just ensuring it's here)
        final_metadatas = []
        for i, meta in enumerate(metadatas_list):
            full_meta = meta.copy()
            full_meta.update(
                {
                    "source_uri": source_uri,
                    "project_id": project_id,
                    "is_global": is_global,
                    "original_text": chunks[i],  # Store original chunk text in metadata
                    "chunk_id": ids_list[i],
                }
            )
            final_metadatas.append(full_meta)

        # --- Sanitize metadata for ChromaDB (remove None values) ---
        sanitized_metadatas = []
        for meta_dict in final_metadatas:
            # Create a new dict excluding keys with None values
            sanitized_dict = {k: v for k, v in meta_dict.items() if v is not None}
            # Ensure all values are str, int, float, or bool for remaining items
            for k, v in list(
                sanitized_dict.items()
            ):  # Iterate over a copy for modification
                if not isinstance(v, (str, int, float, bool)):
                    logger.warning(
                        f"Metadata key '{k}' has unsupported type {type(v)} in chunk {sanitized_dict.get('chunk_id', '?')}. Converting to string."
                    )
                    sanitized_dict[k] = str(v)  # Convert other types to string
                elif (
                    isinstance(v, str) and not v.strip()
                ):  # O3: Remove empty strings too?
                    logger.debug(
                        f"Metadata key '{k}' is an empty or whitespace-only string. Removing for Chroma compatibility."
                    )
                    del sanitized_dict[k]

            sanitized_metadatas.append(sanitized_dict)
        # --------------------------------------------------------

        # Add to Chroma or FAISS
        # --------------------------------------------------------
        # Add to Chroma **or** FAISS
        try:
            # ──────────────────────────────
            # 1) CHROMA branch
            # ──────────────────────────────
            if self.vector_db_type == "chroma":
                if (
                    not self.collection
                    or self.collection.name != self.active_collection_name
                ):
                    logger.warning(
                        "Chroma collection changed or not set. Resetting to '%s'.",
                        self.active_collection_name,
                    )
                    await self.set_collection(self.active_collection_name)

                if not self.collection:
                    logger.error(
                        "Chroma collection '%s' unavailable after reset – aborting add.",
                        self.active_collection_name,
                    )
                    return

                # If we've reached here, self.collection is valid and set.
                # Proceed to add documents.
                self.collection.add(
                    embeddings=embeddings,
                    documents=chunks,
                    metadatas=sanitized_metadatas,
                    ids=ids_list,
                )
                # Log count after successful add
                try:
                    count_after_add = self.collection.count()
                    logger.info(
                        "Chroma collection '%s' new count: %s",
                        self.active_collection_name,
                        count_after_add,
                    )
                except Exception:
                    logger.debug("Could not fetch Chroma count (non-fatal).")

            # ──────────────────────────────
            # 2) FAISS branch
            # ──────────────────────────────
            elif self.vector_db_type == "faiss":
                embeddings_np = np.asarray(embeddings, dtype=np.float32)

                # numeric ids FAISS cares about
                start_id = self.index.ntotal
                new_numeric_ids = list(range(start_id, start_id + len(chunks)))

                # add to index
                self.index.add(embeddings_np)

                # mirror in simple Python docstore
                for i, numeric_id in enumerate(new_numeric_ids):
                    self.docstore["ids"].append(ids_list[i])
                    self.docstore["documents"].append(
                        {
                            "text": chunks[i],
                            "metadata": sanitized_metadatas[i],
                        }
                    )

                logger.debug(
                    "Added %s vectors to FAISS (total now %s).",
                    len(chunks),
                    self.index.ntotal,
                )

                # TODO: persist to disk periodically or on shutdown
                # await self._save_faiss_store()

            # ──────────────────────────────
            # 3) Unsupported backend
            # ──────────────────────────────
            else:
                logger.error(
                    "Unsupported vector_db_type '%s' in _add_chunks_to_vectorstore",
                    self.vector_db_type,
                )
                return

        except Exception as e:
            logger.error(
                "Error while adding chunks to '%s' vector store: %s",
                self.vector_db_type,
                e,
                exc_info=True,
            )
            return
        # ─────────────────────────────────────────────────────────
        # SUCCESS paths continue below … (BM25 update etc.)
        # ─────────────────────────────────────────────────────────
        # --- Update BM25 Index ---------------------------------------------
        try:
            # initialise per-collection containers
            if self.active_collection_name not in self.bm25_corpus:
                self.bm25_corpus[self.active_collection_name] = []
                self.bm25_doc_ids[self.active_collection_name] = []

            # extend corpus and id lists
            self.bm25_corpus[self.active_collection_name].extend(chunks)
            self.bm25_doc_ids[self.active_collection_name].extend(ids_list)

            # rebuild index
            tokenised = [
                doc.lower().split()
                for doc in self.bm25_corpus[self.active_collection_name]
            ]
            self.bm25_indices[self.active_collection_name] = BM25Okapi(tokenised)

            logger.debug(
                "BM25 index for '%s' rebuilt (size=%d)",
                self.active_collection_name,
                len(self.bm25_corpus[self.active_collection_name]),
            )

            # fire-and-forget save to disk
            asyncio.create_task(self._save_bm25_data(self.active_collection_name))

        except Exception as bm_err:
            logger.error(
                "BM25 rebuild failed for '%s': %s",
                self.active_collection_name,
                bm_err,
                exc_info=True,
            )
            # ---------------------------------------------------------------------

            logger.error(
                f"Error during adding chunks to vector store '{self.active_collection_name}': {e}",
                exc_info=True,
            )
        # Don't log semantic_results_raw here
        # logger.debug(
        #     f"Semantic search found {len(semantic_results_raw)} raw candidates from '{self.active_collection_name}'."
        # )

    async def embed_documents(self, documents: List[str]) -> List[List[float]]:
        """Embed multiple documents using the configured embedding manager."""
        if not self._doc_embedding_fn:
            logger.error(
                "Document embedding function attribute (_doc_embedding_fn) not initialized."
            )
            self._ensure_embeddings_initialized()  # Try to initialize
            if not self._doc_embedding_fn:
                raise RuntimeError(
                    "Document embedding function attribute still not available after re-check."
                )

        # self._doc_embedding_fn is synchronous, so run in a thread
        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(None, self._doc_embedding_fn, documents)

    async def embed_query(self, query: str) -> List[float]:
        """Embed a single query using the configured embedding manager. Returns List[float]."""
        if not self._query_embedding_fn:
            logger.error(
                "Query embedding function attribute (_query_embedding_fn) not initialized."
            )
            self._ensure_embeddings_initialized()
            if not self._query_embedding_fn:
                raise RuntimeError(
                    "Query embedding function attribute still not available after re-check."
                )

        # self._query_embedding_fn is synchronous and returns List[List[float]]
        loop = asyncio.get_running_loop()
        embedding_list_of_list = await loop.run_in_executor(
            None, self._query_embedding_fn, query
        )

        if (
            embedding_list_of_list
            and isinstance(embedding_list_of_list, list)
            and len(embedding_list_of_list) == 1
            and isinstance(embedding_list_of_list[0], list)
        ):
            return embedding_list_of_list[0]  # Return the single List[float]
        else:
            logger.error(
                f"Unexpected result from query embedding function. Expected List[List[float]] with one inner list, got: {embedding_list_of_list}"
            )
            # Attempt to handle if it's already List[float] (e.g. from local_fallback if logic changed)
            if (
                isinstance(embedding_list_of_list, list)
                and embedding_list_of_list
                and isinstance(embedding_list_of_list[0], float)
            ):
                logger.warning(
                    "Query embedding function returned List[float] directly. Using as is."
                )
                return embedding_list_of_list
            raise ValueError(
                "Query embedding function returned unexpected result structure."
            )

    async def _save_bm25_data(self, collection_name: str):
        """Saves the BM25 corpus and document IDs for a collection to disk."""
        if collection_name not in self.bm25_corpus:
            logger.warning(
                f"No BM25 corpus data to save for collection: {collection_name}"
            )
            return

        corpus_file = self.bm25_storage_dir / f"{collection_name}_corpus.pkl"
        doc_ids_file = self.bm25_storage_dir / f"{collection_name}_doc_ids.pkl"

        # Data to save
        corpus_to_save = self.bm25_corpus.get(collection_name, [])
        doc_ids_to_save = self.bm25_doc_ids.get(collection_name, [])

        # Function to run in thread
        def _save_sync():
            try:
                logger.info(
                    f"Saving BM25 data for '{collection_name}' to {self.bm25_storage_dir}"
                )
                with open(corpus_file, "wb") as f:
                    pickle.dump(corpus_to_save, f)
                with open(doc_ids_file, "wb") as f:
                    pickle.dump(doc_ids_to_save, f)
                logger.info(f"BM25 data saved successfully for '{collection_name}'.")
            except Exception as e:
                logger.error(
                    f"Error saving BM25 data for '{collection_name}': {e}",
                    exc_info=True,
                )

        await asyncio.to_thread(_save_sync)

    async def _load_bm25_data(self, collection_name: str) -> bool:
        """Loads BM25 corpus and doc IDs, then builds the index if corpus is not empty."""
        corpus_file = self.bm25_storage_dir / f"{collection_name}_corpus.pkl"
        doc_ids_file = self.bm25_storage_dir / f"{collection_name}_doc_ids.pkl"

        def _load_sync():
            try:
                if not corpus_file.exists() or not doc_ids_file.exists():
                    logger.warning(
                        f"BM25 data files not found for collection: {collection_name}. Initializing empty."
                    )
                    self.bm25_corpus[collection_name] = []
                    self.bm25_doc_ids[collection_name] = []
                    self.bm25_indices[collection_name] = (
                        None  # REVERTED: Store None sentinel
                    )
                    return False  # Indicate data wasn't loaded

                logger.info(
                    f"Loading BM25 data for '{collection_name}' from {self.bm25_storage_dir}"
                )
                with open(corpus_file, "rb") as f:
                    corpus = pickle.load(f)
                with open(doc_ids_file, "rb") as f:
                    doc_ids = pickle.load(f)

                self.bm25_corpus[collection_name] = corpus
                self.bm25_doc_ids[collection_name] = doc_ids

                # --- FIX: Only build index if corpus is not empty ---
                if corpus:
                    logger.info(
                        f"Tokenizing corpus and building BM25 index for '{collection_name}' (size: {len(corpus)})..."
                    )
                    start_time = time.time()
                    tokenized_corpus = [doc.lower().split(" ") for doc in corpus]
                    # Create the BM25Okapi object here
                    self.bm25_indices[collection_name] = BM25Okapi(tokenized_corpus)
                    end_time = time.time()
                    logger.info(
                        f"BM25 index for '{collection_name}' built in {end_time - start_time:.2f}s."
                    )
                else:
                    logger.warning(
                        f"Loaded empty BM25 corpus for '{collection_name}'. Index will be None."
                    )
                    # Store None sentinel if corpus is empty after loading
                    self.bm25_indices[collection_name] = (
                        None  # REVERTED: Store None sentinel
                    )

                return True  # Indicate successful load attempt
            except Exception as e:
                logger.error(
                    f"Error loading BM25 data for '{collection_name}': {e}. Initializing empty.",
                    exc_info=True,
                )
                self.bm25_corpus[collection_name] = []
                self.bm25_doc_ids[collection_name] = []
                self.bm25_indices[collection_name] = (
                    None  # REVERTED: Store None sentinel on error too
                )
                return False

        return await asyncio.to_thread(_load_sync)

    async def retrieve_documents(
        self,
        query: str,
        collection_name: Optional[str] = None,
        project_id: Optional[str] = None,
        limit: int = 10,
        filters: Optional[Dict[str, Any]] = None,
        include_global: bool = False,  # New flag to include global documents
        retrieval_mode: str = "semantic",  # "semantic", "bm25", or "hybrid"
        bm25_weight: float = 0.5,  # Weight for BM25 in hybrid mode
    ) -> List[Dict[str, Any]]:
        """
        Retrieve relevant document chunks from the specified collection.

        Args:
            query: The query string.
            collection_name: The name of the collection to query. Defaults to active collection.
            project_id: The project_id to filter by (for project-specific collections).
            limit: The maximum number of documents to return.
            filters: Additional metadata filters for the query.
            include_global: If True, also search in the global collection if project_id is specified.
            retrieval_mode: "semantic", "bm25", or "hybrid".
            bm25_weight: Weight for BM25 score in hybrid retrieval (0 to 1).

        Returns:
            A list of dictionaries, where each dictionary represents a retrieved chunk
            and includes 'page_content', 'metadata', and 'score'.
        """
        if not query:
            logger.warning("Retrieve_documents called with an empty query.")
            return []

        target_collection_name = collection_name or self.active_collection_name
        if not target_collection_name:
            logger.error(
                "No collection specified and no active collection set for document retrieval."
            )
            return []

        # Ensure the target collection is active
        if self.active_collection_name != target_collection_name:
            await self.set_collection(target_collection_name)
        elif self.vector_db_type == "chroma" and (
            not self.collection or self.collection.name != target_collection_name
        ):
            # Re-ensure collection object is valid if name matches but object might be stale
            await self.set_collection(target_collection_name)

            logger.info(
                f"Retrieving documents for query: '{query[:50]}...' from collection: '{target_collection_name}', "
                f"limit: {limit}, project_id: {project_id}, include_global: {include_global}, "
                f"filters: {filters}, mode: {retrieval_mode}"
            )

        semantic_results = []
        bm25_results = []

        # --- Semantic Search ---
        if retrieval_mode in ["semantic", "hybrid"]:
            logger.debug(f"Performing semantic search part for query: {query}")
            try:
                # Get the query embedding function (sync or async)
                embed_query_func = self._query_embedding_fn
                if not embed_query_func:
                    logger.error("Failed to get query embedding function from manager.")
                    return []

                # Execute the function (await if async, run in thread if sync)
                if asyncio.iscoroutinefunction(embed_query_func):
                    raw_embedding_result = await embed_query_func(query)
                else:
                    raw_embedding_result = await asyncio.to_thread(
                        embed_query_func, query
                    )

                # Ensure query_embedding is the actual vector, not nested list
                if (
                    isinstance(raw_embedding_result, list)
                    and len(raw_embedding_result) == 1
                    and isinstance(raw_embedding_result[0], list)
                ):
                    query_embedding = raw_embedding_result[0]
                    logger.debug(
                        "Adjusted nested embedding list from embedding function."
                    )
                elif (
                    isinstance(raw_embedding_result, list)
                    and len(raw_embedding_result) > 0
                    and isinstance(raw_embedding_result[0], float)
                ):
                    query_embedding = raw_embedding_result
                else:
                    logger.error(
                        f"Unexpected embedding result format: {type(raw_embedding_result)}"
                    )
                    return []

                if not query_embedding or not isinstance(query_embedding, list):
                    logger.error(
                        f"Failed to generate a valid query embedding for: {query}"
                    )
                    return []
            except Exception as e:
                logger.error(f"Error generating query embedding: {e}", exc_info=True)
                return []

            if self.vector_db_type == "chroma":
                if not self.collection:
                    logger.error(
                        f"Chroma collection '{target_collection_name}' is not available for query."
                    )
                    return []

                # --- O3 Playbook 10 Prep: Separate ID filter from metadata filter ---
                doc_id_filter = None  # Explicitly keep None for where_document unless used for text search
                metadata_filters = filters.copy() if filters else {}

                # Build the project/global $or filter (for metadata -> where)
                project_or_global_meta_filters = []
                if project_id:
                    project_or_global_meta_filters.append({"project_id": project_id})
                if include_global:
                    # Check for project_id is None OR is_global is True
                    project_or_global_meta_filters.append(
                        {"project_id": "_none_"}
                    )  # Use the sanitized None value
                    project_or_global_meta_filters.append({"is_global": True})

                # Build the final WHERE filter (for metadata)
                where_clauses = []
                if project_or_global_meta_filters:
                    unique_pg_filters = [
                        dict(t)
                        for t in {
                            tuple(d.items()) for d in project_or_global_meta_filters
                        }
                    ]
                    if len(unique_pg_filters) == 1:
                        where_clauses.append(unique_pg_filters[0])
                    elif len(unique_pg_filters) > 1:
                        where_clauses.append({"$or": unique_pg_filters})

                if (
                    metadata_filters
                ):  # Add remaining filters from the original 'filters' arg
                    if isinstance(metadata_filters, dict) and metadata_filters:
                        where_clauses.append(metadata_filters)
                    else:
                        logger.warning(
                            f"Received invalid or empty remaining metadata filters: {metadata_filters}. Ignoring."
                        )

                final_where_filter = None  # For metadata
                if len(where_clauses) == 1:
                    final_where_filter = where_clauses[0]
                elif len(where_clauses) > 1:
                    final_where_filter = {"$and": where_clauses}

                logger.debug(
                    f"Chroma final_where_filter (metadata): {json.dumps(final_where_filter, default=str)}"
                )
                # logger.debug( # doc_id_filter is now always None for this fix
                #     f"Chroma final_where_document_filter (doc_id): {json.dumps(doc_id_filter, default=str)}"
                # )

                try:
                    results = self.collection.query(
                        query_embeddings=[query_embedding],
                        n_results=limit * 2,  # Increased limit for hybrid later
                        where=final_where_filter,  # For metadata
                        where_document=doc_id_filter,  # Should be None if doc_id is metadata
                        include=["metadatas", "documents", "distances"],
                    )
                except Exception as chroma_exc:
                    logger.error("ChromaDB query failed: %s", chroma_exc, exc_info=True)
                    logger.error(
                        "Failed query filter was: %s",
                        json.dumps(final_where_filter, default=str),
                    )
                    return []

                if results and results.get("ids") and results["ids"][0]:
                    for i, doc_id_str in enumerate(results["ids"][0]):
                        doc_content = (
                            results["documents"][0][i]
                            if results["documents"] and results["documents"][0]
                            else ""
                        )
                        metadata = (
                            results["metadatas"][0][i]
                            if results["metadatas"] and results["metadatas"][0]
                            else {}
                        )
                        distance = (
                            results["distances"][0][i]
                            if results["distances"] and results["distances"][0]
                            else float("inf")
                        )
                        # Convert distance to similarity score (e.g., 1 - distance for cosine)
                        score = 1 - distance if distance != float("inf") else 0
                        semantic_results.append(
                            {
                                "id": doc_id_str,
                                "page_content": doc_content,
                                "metadata": metadata,
                                "score": score,
                                "retrieval_type": "semantic",
                            }
                        )
                logger.debug(f"Semantic search found {len(semantic_results)} results.")

            elif self.vector_db_type == "faiss":
                # FAISS retrieval logic (assuming self.faiss_index and self.faiss_docstore are loaded)
                if not self.faiss_index or not self.faiss_docstore:
                    logger.error(
                        "FAISS index or docstore not initialized for retrieval."
                    )
                    return []
                try:
                    query_embedding_np = np.array([query_embedding], dtype=np.float32)
                    distances, indices = self.faiss_index.search(
                        query_embedding_np, limit * 2
                    )

                    for i, idx in enumerate(indices[0]):
                        if idx == -1:  # Faiss uses -1 for no result
                            continue
                        doc_id_str = self.faiss_docstore.get_doc_id(
                            idx
                        )  # Assuming docstore has a way to get original ID
                        doc_content = self.faiss_docstore.get_document(
                            idx
                        )  # Assuming this method exists
                        metadata = self.faiss_docstore.get_metadata(
                            idx
                        )  # Assuming this method exists
                        score = (
                            1 - distances[0][i]
                        )  # Convert L2 distance to similarity if that's what Faiss returns

                        # Apply filters manually for FAISS if 'filters' or 'project_id' is provided
                        passes_filter = True
                        if (
                            final_where_filter
                        ):  # FAISS doesn't support 'where' directly, so manual filter
                            # This is a simplified representation; complex $or/$and requires more logic
                            # For now, assuming filters is a simple dict of key-value pairs
                            if isinstance(final_where_filter, dict):
                                for key, value in final_where_filter.items():
                                    if metadata.get(key) != value:
                                        passes_filter = False

                        if passes_filter:
                            semantic_results.append(
                                {
                                    "id": doc_id_str,
                                    "page_content": doc_content,
                                    "metadata": metadata,
                                    "score": score,
                                    "retrieval_type": "semantic",
                                }
                            )
                    semantic_results = semantic_results[
                        :limit
                    ]  # Apply limit after filtering
                    logger.debug(
                        f"FAISS semantic search found {len(semantic_results)} results after filtering."
                    )
                except Exception as faiss_exc:
                    logger.error(f"FAISS search failed: {faiss_exc}", exc_info=True)
                    return []

        # --- BM25 Search ---
        if retrieval_mode in ["bm25", "hybrid"]:
            logger.debug(f"Performing BM25 search part for query: {query}")
            # --- FIX: Check if index is None (empty or failed load) ---
            bm25_index = self.bm25_indices.get(target_collection_name)
            if bm25_index is None:
                logger.warning(
                    f"BM25 index for '{target_collection_name}' is empty or not available. Skipping BM25 part."
                )
            else:
                tokenized_query = query.lower().split()
                doc_scores = bm25_index.get_scores(tokenized_query)

                # Get top N document indices
                top_n_indices = np.argsort(doc_scores)[::-1][
                    : limit * 2
                ]  # Fetch more for hybrid

                for i in top_n_indices:
                    # Filter by project_id and general filters if provided
                    doc_id = self.bm25_doc_ids[target_collection_name][i]
                    # We need to fetch the actual document content and metadata.
                    # This requires a lookup, perhaps from Chroma/FAISS by ID or a separate store.
                    # For simplicity, we assume we can get it from Chroma if it exists there.
                    # This part is tricky as BM25 operates on its own corpus.
                    # A common approach is to use BM25 to get IDs, then fetch from primary vector store.

                    # Placeholder: fetch from Chroma by ID if available
                    try:
                        if self.collection:  # Chroma
                            fetched_doc = self.collection.get(
                                ids=[doc_id], include=["metadatas", "documents"]
                            )
                            if fetched_doc and fetched_doc["ids"]:
                                metadata = (
                                    fetched_doc["metadatas"][0]
                                    if fetched_doc["metadatas"]
                                    else {}
                                )
                                content = (
                                    fetched_doc["documents"][0]
                                    if fetched_doc["documents"]
                                    else ""
                                )

                                # Apply filters manually
                                passes_filter = True
                                current_project_id = metadata.get("project_id")
                                is_doc_global = metadata.get("is_global", False)

                                # Project/Global filter logic
                                project_match = False
                                if project_id and current_project_id == project_id:
                                    project_match = True
                                if include_global and (
                                    current_project_id is None or is_doc_global
                                ):
                                    project_match = True
                                if (
                                    project_id and not project_match
                                ):  # If project_id given, but no match
                                    passes_filter = False

                                if passes_filter and filters:
                                    for key, value in filters.items():
                                        if metadata.get(key) != value:
                                            passes_filter = False
                                            break

                                if passes_filter:
                                    bm25_results.append(
                                        {
                                            "id": doc_id,
                                            "page_content": content,
                                            "metadata": metadata,
                                            "score": doc_scores[i],  # Raw BM25 score
                                            "retrieval_type": "bm25",
                                        }
                                    )
                        # Add FAISS equivalent if BM25 is used with FAISS
                    except Exception as e:
                        logger.warning(
                            f"Could not fetch doc {doc_id} for BM25 result: {e}"
                        )

                logger.debug(
                    f"BM25 search found {len(bm25_results)} candidate results before filtering."
                )
                bm25_results = bm25_results[:limit]  # Ensure limit is applied

        # --- Hybrid Search (Combine and Re-rank) ---
        combined_results = []
        if retrieval_mode == "hybrid":
            # Combine semantic and BM25 results, normalize scores, and re-rank or sort
            # Simple weighted sum for now, assuming scores are somewhat comparable (needs normalization)

            # Normalize BM25 scores (e.g., min-max or just ensure positive)
            # For now, let's assume BM25 scores are positive. Higher is better.
            # Semantic scores are already 0-1 (similarity).

            temp_hybrid_results = {}  # Use dict to handle duplicates by ID

            for res in semantic_results:
                res_id = res["id"]
                if res_id not in temp_hybrid_results:
                    temp_hybrid_results[res_id] = res
                    temp_hybrid_results[res_id]["normalized_score"] = res["score"] * (
                        1 - bm25_weight
                    )
                else:  # If somehow present, add scores (less likely with UUIDs)
                    temp_hybrid_results[res_id]["normalized_score"] += res["score"] * (
                        1 - bm25_weight
                    )

            max_bm25_score = (
                max(r["score"] for r in bm25_results) if bm25_results else 1.0
            )
            min_bm25_score = (
                min(r["score"] for r in bm25_results) if bm25_results else 0.0
            )

            for res in bm25_results:
                res_id = res["id"]
                # Normalize BM25 score to ~0-1 range (simple min-max)
                norm_bm25_score = 0
                if max_bm25_score > min_bm25_score:  # Avoid division by zero
                    norm_bm25_score = (res["score"] - min_bm25_score) / (
                        max_bm25_score - min_bm25_score
                    )
                elif max_bm25_score > 0:  # All scores are same and positive
                    norm_bm25_score = 1.0

                weighted_bm25_score = norm_bm25_score * bm25_weight

                if res_id not in temp_hybrid_results:
                    temp_hybrid_results[res_id] = res
                    temp_hybrid_results[res_id][
                        "normalized_score"
                    ] = weighted_bm25_score
                else:
                    temp_hybrid_results[res_id][
                        "normalized_score"
                    ] += weighted_bm25_score
                    # Merge metadata if needed, preferring semantic if conflict (or define priority)
                    temp_hybrid_results[res_id]["metadata"].update(res["metadata"])
                    # Ensure page_content is consistent if merging
                    if not temp_hybrid_results[res_id].get("page_content") and res.get(
                        "page_content"
                    ):
                        temp_hybrid_results[res_id]["page_content"] = res.get(
                            "page_content"
                        )

            # Sort by the combined normalized score
            combined_results = sorted(
                temp_hybrid_results.values(),
                key=lambda x: x.get("normalized_score", 0),
                reverse=True,
            )
            logger.info(
                f"Hybrid search combined {len(semantic_results)} semantic and {len(bm25_results)} BM25 results into {len(combined_results)} results."
            )

        elif retrieval_mode == "semantic":
            combined_results = sorted(
                semantic_results, key=lambda x: x.get("score", 0), reverse=True
            )
        elif retrieval_mode == "bm25":
            combined_results = sorted(
                bm25_results, key=lambda x: x.get("score", 0), reverse=True
            )

        return combined_results[:limit]

    async def _update_bm25_index(
        self, chunks: List[Dict[str, Any]], collection_name: str
    ):
        if not chunks:
            return

        try:
            logger.info(f"Updating BM25 index for collection '{collection_name}'...")
            new_texts = [chunk["text"] for chunk in chunks]
            new_doc_ids = [chunk["id"] for chunk in chunks]

            if collection_name not in self.bm25_corpus:
                await self._load_bm25_data(collection_name)

            corpus = self.bm25_corpus.setdefault(collection_name, [])
            doc_ids = self.bm25_doc_ids.setdefault(collection_name, [])

            corpus.extend(new_texts)
            doc_ids.extend(new_doc_ids)

            tokenized_corpus = [doc.lower().split(" ") for doc in corpus]
            self.bm25_indices[collection_name] = BM25Okapi(tokenized_corpus)
            logger.info(
                f"BM25 index for '{collection_name}' updated. Corpus size: {len(corpus)}"
            )

            await self._save_bm25_data(collection_name)
        except Exception as bm_err:
            logger.error(
                f"Error updating BM25 index for {collection_name}: {bm_err}",
                exc_info=True,
            )
            # Decide if failure here is critical - currently just logs

    def get_bm25_retrieval_data_for_collection(
        self, collection_name: str
    ) -> Optional[Tuple[List[str], List[List[str]], BM25Okapi]]:
        """
        Returns the data needed for HybridRetriever's BM25 component for a given collection.
        Ensures BM25 index is loaded/built.

        Returns:
            A tuple (doc_ids, tokenized_corpus, bm25_okapi_instance) or None if data is unavailable.
        """
        if (
            collection_name not in self.bm25_indices
            or self.bm25_indices[collection_name] is None
        ):
            logger.warning(
                f"BM25 index for '{collection_name}' not available or not built. Cannot provide BM25 data."
            )
            # Optionally, attempt to load/build it here if that's desired behavior
            # asyncio.run(self._load_bm25_data(collection_name)) # Careful with running async from sync
            return None

        doc_ids = self.bm25_doc_ids.get(collection_name, [])
        corpus_texts = self.bm25_corpus.get(
            collection_name, []
        )  # These are the original texts for BM25
        bm25_instance = self.bm25_indices[collection_name]

        if not doc_ids or not corpus_texts or not bm25_instance:
            logger.warning(
                f"Missing BM25 components (ids, corpus, or index object) for '{collection_name}'."
            )
            return None

        # The HybridRetriever wants the full untokenized corpus texts as well.
        # And it builds its own BM25Okapi from a tokenized version.
        # So, we need to provide the untokenized corpus texts and the doc_ids.
        # The HybridRetriever also needs a tokenized version for its internal BM25. Let's provide that too.
        tokenized_corpus_for_hybrid = [doc.lower().split(" ") for doc in corpus_texts]

        return doc_ids, corpus_texts, tokenized_corpus_for_hybrid, bm25_instance

    # ... (get_document_count, clear_vectorstore, process_and_ingest_pdf_from_path, _save_faiss_store) ...

    async def index_document_content(
        self,
        doc_id: str,
        content: str,
        doc_metadata: Dict[str, Any],
    ) -> bool:
        """
        Index in-memory document content. Chunks the content and adds to vector store.

        Args:
            doc_id: The unique identifier for the document.
            content: The textual content of the document.
            doc_metadata: A dictionary of metadata for the document.
                          Expected to contain 'project_id', 'document_type', 'source', etc.

        Returns:
            True if indexing was successful, False otherwise.
        """
        logger.info(
            f"Starting in-memory content indexing for doc_id: {doc_id}, type: {doc_metadata.get('document_type')}"
        )

        try:
            # Ensure essential metadata is present
            if not all(
                k in doc_metadata for k in ["project_id", "document_type", "source"]
            ):
                logger.warning(
                    f"Missing essential metadata (project_id, document_type, or source) for doc_id: {doc_id}. Using defaults or skipping."
                )
                # Ensure defaults if possible, or handle error
                doc_metadata.setdefault(
                    "source", doc_id
                )  # Default source to doc_id if missing
                if (
                    "document_type" not in doc_metadata
                ):  # Cannot proceed without document_type for collection policy
                    logger.error(
                        f"Cannot determine collection for doc_id: {doc_id} due to missing document_type."
                    )
                    return False

            target_collection_name = self._get_collection_name_for_doc(doc_metadata)
            if not target_collection_name:
                logger.error(
                    f"Collection policy failed to determine a target collection name for doc_id: {doc_id}."
                )
                return False

            logger.info(
                f"Target collection for doc_id {doc_id}: {target_collection_name}"
            )

            needs_to_set_collection = (
                self.active_collection_name != target_collection_name
            ) or (
                self.vector_db_type == "chroma"
                and (
                    not self.collection
                    or self.collection.name != target_collection_name
                )
            )
            if needs_to_set_collection:
                logger.info(
                    f"Setting active collection to '{target_collection_name}' for content indexing."
                )
                await self.set_collection(target_collection_name)

            # Chunk the provided content
            # _chunk_text expects 'metadata' to be the base doc metadata
            chunk_data_list = await self._chunk_text(
                text=content, document_id=doc_id, metadata=doc_metadata
            )

            if not chunk_data_list:
                logger.warning(f"No chunks created from content for document {doc_id}.")
                return False

            chunks_texts = [item["text"] for item in chunk_data_list]
            metadatas_list = [item["metadata"] for item in chunk_data_list]
            ids_list = [item["id"] for item in chunk_data_list]

            # Add to vector store
            await self._add_chunks_to_vectorstore(
                chunks=chunks_texts,
                metadatas_list=metadatas_list,
                ids_list=ids_list,
                source_uri=doc_metadata.get(
                    "source"
                ),  # 'source' should be in doc_metadata
                project_id=doc_metadata.get("project_id"),
                is_global=doc_metadata.get("is_global", False),
            )

            logger.info(
                f"Successfully processed and initiated addition of {len(chunks_texts)} chunks from in-memory content for doc_id '{doc_id}' to collection '{self.active_collection_name}'."
            )
            return True

        except Exception as e_exc:
            logger.error(
                f"Error indexing in-memory content for doc_id {doc_id}: {e_exc}",
                exc_info=True,
            )
            return False
